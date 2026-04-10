import os
import uuid
from datetime import datetime, timedelta
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, current_app, abort
from flask_login import login_required, current_user
from sqlalchemy import or_, and_
import tempfile
from werkzeug.utils import secure_filename
from flask import jsonify
from sqlalchemy import func, desc
import io, csv

from app.admin.forms import (DebtorForm, InteractionForm, ImportForm, UserForm, 
                            EditUserForm, ChangePasswordForm, DocumentUploadForm)

from app import db
from app.models import Debtor, Interaction, User, Document
from app.admin.utils import import_excel

bp = Blueprint('admin', __name__, url_prefix='/admin')


def admin_required(f):
    """Decorator to require admin role."""
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.can_manage_users():
            flash('У вас нет доступа к этой странице', 'danger')
            return redirect(url_for('admin.dashboard'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function


def editor_required(f):
    """Decorator to require edit permissions (admin or manager)."""
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.can_edit_debtors():
            flash('У вас нет прав для редактирования', 'danger')
            return redirect(url_for('admin.dashboard'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function


@bp.route('/dashboard')
@login_required
def dashboard():
    stats = {
        'total_debtors': Debtor.query.count(),
        'total_debt': db.session.query(db.func.sum(Debtor.debt_amount)).scalar() or 0,
        'recent_interactions': Interaction.query.order_by(Interaction.date_time.desc()).limit(10).all()
    }
    return render_template('admin/dashboard.html', stats=stats)


@bp.route('/debtors')
@login_required
def debtors():
    """
    Список должников с возможностью быстрых фильтров:
    + добавлен фильтр status (active / archived / paid)
    """
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    bank_filter = request.args.get('bank', '', type=str)

    # Новые параметры быстрых фильтров
    min_debt_raw = request.args.get('min_debt', '', type=str)
    max_debt_raw = request.args.get('max_debt', '', type=str)
    date_from_raw = request.args.get('date_from', '', type=str)
    date_to_raw = request.args.get('date_to', '', type=str)
    sort = request.args.get('sort', '', type=str)
    status_filter = request.args.get('status', '', type=str)  # <-- новый параметр

    # Попытаемся корректно распарсить суммы и даты
    min_debt = None
    max_debt = None
    date_from = None
    date_to = None

    try:
        if min_debt_raw:
            min_debt = float(min_debt_raw)
    except ValueError:
        min_debt = None

    try:
        if max_debt_raw:
            max_debt = float(max_debt_raw)
    except ValueError:
        max_debt = None

    try:
        if date_from_raw:
            # ожидание формата YYYY-MM-DD (html input date)
            date_from = datetime.strptime(date_from_raw, '%Y-%m-%d')
    except ValueError:
        date_from = None

    try:
        if date_to_raw:
            # включительно: до конца дня
            parsed = datetime.strptime(date_to_raw, '%Y-%m-%d')
            date_to = parsed.replace(hour=23, minute=59, second=59)
    except ValueError:
        date_to = None

    query = Debtor.query

    # Apply search
    if search:
        query = query.filter(
            or_(
                Debtor.full_name.ilike(f'%{search}%'),
                Debtor.contract_number.ilike(f'%{search}%'),
                Debtor.bank_name.ilike(f'%{search}%')
            )
        )

    # Apply bank filter
    if bank_filter:
        query = query.filter(Debtor.bank_name == bank_filter)

    # Apply status filter (active / archived / paid)
    if status_filter:
        query = query.filter(Debtor.status == status_filter)

    # Apply debt amount filters
    if min_debt is not None:
        query = query.filter(Debtor.debt_amount >= min_debt)
    if max_debt is not None:
        query = query.filter(Debtor.debt_amount <= max_debt)

    # Apply date filters (по created_at)
    if date_from is not None:
        query = query.filter(Debtor.created_at >= date_from)
    if date_to is not None:
        query = query.filter(Debtor.created_at <= date_to)

    # Sorting
    if sort == 'debt_asc':
        order_by = Debtor.debt_amount.asc()
    elif sort == 'debt_desc':
        order_by = Debtor.debt_amount.desc()
    elif sort == 'date_asc':
        order_by = Debtor.created_at.asc()
    elif sort == 'date_desc':
        order_by = Debtor.created_at.desc()
    else:
        order_by = Debtor.created_at.desc()

    # Get unique banks for filter dropdown
    banks = db.session.query(Debtor.bank_name).distinct().order_by(Debtor.bank_name).all()
    banks = [b[0] for b in banks]

    # Paginate
    pagination = query.order_by(order_by).paginate(
        page=page,
        per_page=current_app.config.get('DEBTORS_PER_PAGE', 20),
        error_out=False
    )

    # Подготовим дату для пресета "Последние 30 дней" (формат YYYY-MM-DD)
    date_30 = (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d')

    # Передаём параметры обратно в шаблон, чтобы сохранялись в ссылках/полях
    return render_template('admin/debtors.html',
                           debtors=pagination.items,
                           pagination=pagination,
                           banks=banks,
                           search=search,
                           bank_filter=bank_filter,
                           min_debt=min_debt_raw,
                           max_debt=max_debt_raw,
                           date_from=date_from_raw,
                           date_to=date_to_raw,
                           sort=sort,
                           date_30=date_30,
                           status_filter=status_filter)  # <-- передаю в шаблон



@bp.route('/debtors/<int:debtor_id>')
@login_required
def debtor_detail(debtor_id):
    debtor = Debtor.query.get_or_404(debtor_id)
    form = DocumentUploadForm()
    interactions = Interaction.query.filter_by(debtor_id=debtor_id).order_by(Interaction.date_time.desc()).all()

    # Безопасно получаем MAX_CONTENT_LENGTH; если он не задан, используем 16 MB по умолчанию
    max_bytes = current_app.config.get('MAX_CONTENT_LENGTH') or (16 * 1024 * 1024)
    try:
        max_upload_mb = int(max_bytes) // (1024 * 1024)
    except (TypeError, ValueError):
        max_upload_mb = 16

    return render_template(
        'admin/debtor_detail.html',
        debtor=debtor,
        interactions=interactions,
        form=form,
        max_upload_mb=max_upload_mb
    )

@bp.route('/debtors/add', methods=['GET', 'POST'])
@login_required
@editor_required
def add_debtor():
    form = DebtorForm()
    if form.validate_on_submit():
        # Check if contract number already exists
        existing = Debtor.query.filter_by(contract_number=form.contract_number.data).first()
        if existing:
            flash('Договор с таким номером уже существует', 'danger')
            return redirect(url_for('admin.add_debtor'))
        
        debtor = Debtor(
            full_name=form.full_name.data,
            birth_date=form.birth_date.data,
            passport_data=form.passport_data.data,
            contract_number=form.contract_number.data,
            debt_amount=form.debt_amount.data,
            bank_name=form.bank_name.data,
            status=form.status.data or 'active'
        )
        db.session.add(debtor)
        db.session.commit()
        flash('Должник успешно добавлен', 'success')
        return redirect(url_for('admin.debtor_detail', debtor_id=debtor.id))
    
    return render_template('admin/debtor_form.html', form=form, title='Добавить должника')


@bp.route('/debtors/<int:debtor_id>/edit', methods=['GET', 'POST'])
@login_required
@editor_required
def edit_debtor(debtor_id):
    debtor = Debtor.query.get_or_404(debtor_id)
    form = DebtorForm(obj=debtor)
    
    if form.validate_on_submit():
        # Check if contract number changed and already exists
        if form.contract_number.data != debtor.contract_number:
            existing = Debtor.query.filter_by(contract_number=form.contract_number.data).first()
            if existing:
                flash('Договор с таким номером уже существует', 'danger')
                return redirect(url_for('admin.edit_debtor', debtor_id=debtor_id))
        
        debtor.full_name = form.full_name.data
        debtor.birth_date = form.birth_date.data
        debtor.passport_data = form.passport_data.data
        debtor.contract_number = form.contract_number.data
        debtor.debt_amount = form.debt_amount.data
        debtor.bank_name = form.bank_name.data
        debtor.status = form.status.data or debtor.status
        
        db.session.commit()
        flash('Данные должника обновлены', 'success')
        return redirect(url_for('admin.debtor_detail', debtor_id=debtor.id))
    
    return render_template('admin/debtor_form.html', form=form, title='Редактировать должника', debtor=debtor)


@bp.route('/debtors/<int:debtor_id>/documents/upload', methods=['POST'])
@login_required
@editor_required
def upload_debtor_document(debtor_id):
    debtor = Debtor.query.get_or_404(debtor_id)
    # используем форму DocumentUploadForm
    from app.admin.forms import DocumentUploadForm
    form = DocumentUploadForm()
    if not form.validate_on_submit():
        for field, errors in form.errors.items():
            for err in errors:
                flash(f"{field}: {err}", 'danger')
        return redirect(url_for('admin.debtor_detail', debtor_id=debtor_id))

    file = form.file.data
    if not file or not getattr(file, 'filename', None):
        flash('Файл не выбран', 'danger')
        return redirect(url_for('admin.debtor_detail', debtor_id=debtor_id))

    original_name = secure_filename(file.filename)
    if not original_name:
        flash('Неверное имя файла', 'danger')
        return redirect(url_for('admin.debtor_detail', debtor_id=debtor_id))

    # Проверка расширения
    ext = os.path.splitext(original_name)[1].lower().lstrip('.')
    allowed = current_app.config.get('ALLOWED_UPLOAD_EXTENSIONS',
                                     {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx', 'xls', 'xlsx', 'txt'})
    if ext not in allowed:
        flash('Недопустимый тип файла', 'danger')
        return redirect(url_for('admin.debtor_detail', debtor_id=debtor_id))

    # Безопасный UPLOAD_FOLDER: берем из конфига или используем instance/uploads (fallback)
    upload_folder = current_app.config.get('UPLOAD_FOLDER')
    if not upload_folder:
        # instance_path гарантированно существует у Flask приложения
        upload_folder = os.path.join(current_app.instance_path, 'uploads')

    # Создаём каталог для хран. файлов должника и обеспечиваем права/существование
    debtor_folder = os.path.join(upload_folder, f"debtors_{debtor_id}")
    try:
        os.makedirs(debtor_folder, exist_ok=True)
    except Exception as e:
        current_app.logger.exception("Не удалось создать папку для загрузки: %s", debtor_folder)
        flash('Ошибка сервера: не удалось подготовить папку для загрузки', 'danger')
        return redirect(url_for('admin.debtor_detail', debtor_id=debtor_id))

    # Создаём уникальное имя на диске
    unique_name = f"{uuid.uuid4().hex}_{original_name}"
    disk_path = os.path.join(debtor_folder, unique_name)

    try:
        file.save(disk_path)
        file_size = os.path.getsize(disk_path)
        content_type = getattr(file, 'mimetype', None)
    except Exception as e:
        current_app.logger.exception("Ошибка при сохранении файла: %s", e)
        # Попытка удалить остаточный файл
        try:
            if os.path.exists(disk_path):
                os.remove(disk_path)
        except Exception:
            pass
        flash('Ошибка при сохранении файла', 'danger')
        return redirect(url_for('admin.debtor_detail', debtor_id=debtor_id))

    # Сохраняем запись в БД
    from app.models import Document  # убедимся, что модель доступна
    doc = Document(
        debtor_id=debtor.id,
        filename_original=original_name,
        filename_disk=unique_name,
        content_type=content_type,
        size=file_size,
        uploaded_by=current_user.id
    )
    db.session.add(doc)
    db.session.commit()
    flash('Файл успешно загружен', 'success')
    return redirect(url_for('admin.debtor_detail', debtor_id=debtor_id))

# --- Скачивание документа ---
@bp.route('/debtors/<int:debtor_id>/documents/<int:doc_id>/download')
@login_required
def download_debtor_document(debtor_id, doc_id):
    doc = Document.query.filter_by(id=doc_id, debtor_id=debtor_id).first_or_404()
    upload_folder = current_app.config.get('UPLOAD_FOLDER')
    debtor_folder = os.path.join(upload_folder, f"debtors_{debtor_id}")
    file_path = os.path.join(debtor_folder, doc.filename_disk)
    if not os.path.exists(file_path):
        flash('Файл не найден на сервере', 'danger')
        abort(404)
    # безопасная отправка файла
    return send_file(file_path, as_attachment=True, download_name=doc.filename_original, mimetype=doc.content_type or 'application/octet-stream')

# --- Удаление документа ---
@bp.route('/debtors/<int:debtor_id>/documents/<int:doc_id>/delete', methods=['POST'])
@login_required
@editor_required
def delete_debtor_document(debtor_id, doc_id):
    doc = Document.query.filter_by(id=doc_id, debtor_id=debtor_id).first_or_404()
    # удаляем файл с диска
    upload_folder = current_app.config.get('UPLOAD_FOLDER')
    debtor_folder = os.path.join(upload_folder, f"debtors_{debtor_id}")
    file_path = os.path.join(debtor_folder, doc.filename_disk)
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        current_app.logger.warning(f"Не удалось удалить файл {file_path}: {e}")
    # удаляем запись из БД
    db.session.delete(doc)
    db.session.commit()
    flash('Документ удалён', 'success')
    return redirect(url_for('admin.debtor_detail', debtor_id=debtor_id))

@bp.route('/debtors/<int:debtor_id>/delete', methods=['POST'])
@login_required
@editor_required
def delete_debtor(debtor_id):
    debtor = Debtor.query.get_or_404(debtor_id)
    db.session.delete(debtor)
    db.session.commit()
    flash('Должник удалён', 'success')
    return redirect(url_for('admin.debtors'))


@bp.route('/debtors/<int:debtor_id>/interaction/add', methods=['GET', 'POST'])
@login_required
def add_interaction(debtor_id):
    debtor = Debtor.query.get_or_404(debtor_id)
    form = InteractionForm()
    
    if form.validate_on_submit():
        try:
            # Объединяем дату и время
            date_time_str = f"{form.date.data} {form.time.data}"
            date_time_obj = datetime.strptime(date_time_str, '%d.%m.%Y %H:%M')
            
            interaction = Interaction(
                debtor_id=debtor.id,
                user_id=current_user.id,
                interaction_type=form.interaction_type.data,
                date_time=date_time_obj,
                result=form.result.data,
                comment=form.comment.data
            )
            db.session.add(interaction)
            db.session.commit()
            flash('Взаимодействие добавлено', 'success')
            return redirect(url_for('admin.debtor_detail', debtor_id=debtor.id))
        
        except ValueError:
            flash('Неверный формат даты или времени', 'danger')
    
    # Устанавливаем текущие дату и время по умолчанию
    if not form.date.data:
        form.date.data = datetime.now().strftime('%d.%m.%Y')
    if not form.time.data:
        form.time.data = datetime.now().strftime('%H:%M')
    
    return render_template('admin/interaction_form.html', form=form, debtor=debtor)


@bp.route('/import', methods=['GET', 'POST'])
@login_required
@editor_required
def import_data():
    form = ImportForm()
    
    if form.validate_on_submit():
        # Получаем загруженный файл
        file = form.file.data
        
        if file:
            # Создаем временный файл
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                file.save(tmp_file.name)
                filepath = tmp_file.name
            
            try:
                # Импортируем данные
                stats = import_excel(filepath)
                
                if stats['errors']:
                    flash('Импорт завершён с ошибками:', 'warning')
                    for error in stats['errors'][:5]:  # Показываем только первые 5 ошибок
                        flash(error, 'danger')
                    if len(stats['errors']) > 5:
                        flash(f'... и ещё {len(stats["errors"]) - 5} ошибок', 'danger')
                else:
                    flash(f'Импорт успешен: создано {stats["created"]}, обновлено {stats["updated"]}', 'success')
                
                # Показываем статистику даже при ошибках
                if stats['created'] > 0 or stats['updated'] > 0:
                    flash(f'Статистика: создано {stats["created"]}, обновлено {stats["updated"]}', 'info')
                
            except Exception as e:
                flash(f'Ошибка импорта: {str(e)}', 'danger')
            finally:
                # Удаляем временный файл
                if os.path.exists(filepath):
                    os.remove(filepath)
            
            return redirect(url_for('admin.import_data'))
    
    return render_template('admin/import.html', form=form)


@bp.route('/generate-test-file')
@login_required
@editor_required
def generate_test_file():
    """Генерирует тестовый Excel файл для импорта"""
    try:
        from openpyxl import Workbook
        import tempfile
        import os
        
        # Создаем временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            filepath = tmp_file.name
        
        # Создаем тестовые данные
        wb = Workbook()
        ws = wb.active
        ws.title = "Должники"
        
        # Заголовки
        headers = ['ФИО', 'Дата рождения', 'Паспорт', 'Номер договора', 'Сумма долга', 'Банк']
        ws.append(headers)
        
        # Тестовые данные
        test_data = [
            ['Иванов Иван Иванович', '15.03.1980', '1234 567890', 'TEST-001', 150000, 'Сбербанк'],
            ['Петрова Мария Сергеевна', '22.07.1985', '5678 123456', 'TEST-002', 75500.50, 'ВТБ'],
            ['Сидоров Петр Алексеевич', '03.11.1975', '9012 345678', 'TEST-003', 200000, 'Альфа-Банк'],
            ['Кузнецова Анна Дмитриевна', '18.05.1990', '3456 789012', 'TEST-004', 50000, 'Тинькофф'],
            ['Федоров Алексей Петрович', '29.12.1982', '7890 123456', 'TEST-005', 300000, 'Газпромбанк']
        ]
        
        for row in test_data:
            ws.append(row)
        
        # Настраиваем ширину столбцов
        column_widths = [30, 15, 15, 20, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        wb.save(filepath)
        
        # Отправляем файл пользователю
        return send_file(
            filepath,
            as_attachment=True,
            download_name=f'test_debtors_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Ошибка генерации тестового файла: {str(e)}', 'danger')
        return redirect(url_for('admin.import_data'))


@bp.route('/export')
@login_required
def export_data():
    try:
        from openpyxl import Workbook
        from io import BytesIO

        # Создаем Excel файл в памяти
        wb = Workbook()
        ws = wb.active
        ws.title = "Должники"

        # Заголовки: добавляем Status
        headers = ['ФИО', 'Дата рождения', 'Паспорт', 'Номер договора', 'Сумма долга', 'Банк', 'Статус']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Данные
        debtors = Debtor.query.order_by(Debtor.created_at.desc()).all()
        for row, debtor in enumerate(debtors, 2):
            ws.cell(row=row, column=1, value=debtor.full_name)
            ws.cell(row=row, column=2, value=debtor.birth_date.strftime('%d.%m.%Y') if debtor.birth_date else '')
            ws.cell(row=row, column=3, value=debtor.passport_data)
            ws.cell(row=row, column=4, value=debtor.contract_number)
            ws.cell(row=row, column=5, value=float(debtor.debt_amount))
            ws.cell(row=row, column=6, value=debtor.bank_name)
            # Статус — human readable (Активный/Архив/Погашен)
            status_display = {
                'active': 'Активный',
                'archived': 'Архив',
                'paid': 'Погашен'
            }.get(getattr(debtor, 'status', ''), getattr(debtor, 'status', ''))
            ws.cell(row=row, column=7, value=status_display)

        # Настраиваем ширину столбцов
        column_widths = [30, 15, 15, 20, 15, 20, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Сохраняем в память
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Скачиваем файл
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'debtors_export_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'Ошибка экспорта: {str(e)}', 'danger')
        return redirect(url_for('admin.debtors'))


@bp.route('/users')
@login_required
@admin_required
def users():
    all_users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin/users.html', users=all_users)


@bp.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    form = UserForm()
    
    if form.validate_on_submit():
        # Check if username exists
        existing = User.query.filter_by(username=form.username.data).first()
        if existing:
            flash('Пользователь с таким логином уже существует', 'danger')
            return redirect(url_for('admin.add_user'))
        
        user = User(
            username=form.username.data,
            role=form.role.data
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash(f'Пользователь {user.username} создан', 'success')
        return redirect(url_for('admin.users'))
    
    return render_template('admin/user_form.html', form=form)


@bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Нельзя редактировать самого себя
    if user.id == current_user.id:
        flash('Вы не можете редактировать свой собственный аккаунт', 'warning')
        return redirect(url_for('admin.users'))
    
    form = EditUserForm(obj=user)
    
    if form.validate_on_submit():
        # Проверяем, не занят ли логин другим пользователем
        existing_user = User.query.filter(User.username == form.username.data, User.id != user.id).first()
        if existing_user:
            flash('Пользователь с таким логином уже существует', 'danger')
            return redirect(url_for('admin.edit_user', user_id=user.id))
        
        user.username = form.username.data
        user.role = form.role.data
        db.session.commit()
        
        flash(f'Пользователь {user.username} успешно обновлен', 'success')
        return redirect(url_for('admin.users'))
    
    return render_template('admin/edit_user.html', form=form, user=user)

@bp.route('/users/<int:user_id>/deactivate', methods=['POST'])
@login_required
@admin_required
def deactivate_user(user_id):
    """Деактивация пользователя (мягкое удаление)"""
    user = User.query.get_or_404(user_id)
    
    # Проверки безопасности
    if user.id == current_user.id:
        flash('Вы не можете деактивировать свой собственный аккаунт', 'danger')
        return redirect(url_for('admin.users'))
    
    if not user.is_active:
        flash('Пользователь уже деактивирован', 'warning')
        return redirect(url_for('admin.users'))
    
    username = user.username
    user.is_active = False
    db.session.commit()
    
    flash(f'Пользователь {username} деактивирован', 'success')
    return redirect(url_for('admin.users'))

@bp.route('/users/<int:user_id>/activate', methods=['POST'])
@login_required
@admin_required
def activate_user(user_id):
    """Активация пользователя"""
    user = User.query.get_or_404(user_id)
    
    if user.is_active:
        flash('Пользователь уже активен', 'warning')
        return redirect(url_for('admin.users'))
    
    username = user.username
    user.is_active = True
    db.session.commit()
    
    flash(f'Пользователь {username} активирован', 'success')
    return redirect(url_for('admin.users'))



@bp.route('/users/<int:user_id>/change-password', methods=['GET', 'POST'])
@login_required
@admin_required
def change_user_password(user_id):
    user = User.query.get_or_404(user_id)
    
    # Нельзя менять пароль самому себе
    if user.id == current_user.id:
        flash('Вы не можете изменить пароль своего собственного аккаунта', 'warning')
        return redirect(url_for('admin.users'))
    
    form = ChangePasswordForm()
    
    if form.validate_on_submit():
        user.set_password(form.new_password.data)
        db.session.commit()
        
        flash(f'Пароль пользователя {user.username} успешно изменен', 'success')
        return redirect(url_for('admin.users'))
    
    return render_template('admin/change_password.html', form=form, user=user)

@bp.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    """Полное удаление пользователя (только если нет активности)"""
    user = User.query.get_or_404(user_id)
    
    # Проверки безопасности
    if user.id == current_user.id:
        flash('Вы не можете удалить свой собственный аккаунт', 'danger')
        return redirect(url_for('admin.users'))
    
    # Проверяем, есть ли у пользователя взаимодействия
    # Используем простую проверку наличия связанной истории
    if user.interactions:
        flash(f'Нельзя удалить пользователя {user.username}, так как у него есть история взаимодействий. Используйте деактивацию вместо удаления.', 'warning')
        return redirect(url_for('admin.users'))
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    
    flash(f'Пользователь {username} полностью удален из системы', 'success')
    return redirect(url_for('admin.users'))

# --- Analytics endpoints (append this block at the end of app/admin/routes.py) ---
@bp.route('/analytics')
@login_required
@editor_required
def analytics():
    """Страница аналитики (рендерит шаблон, фронтенд запрашивает /analytics/data)"""
    return render_template('admin/analytics.html')


@bp.route('/analytics/data')
@login_required
@editor_required
def analytics_data():
    """
    Возвращает JSON с аналитическими данными.
    Параметры (GET):
      - start_date (YYYY-MM-DD)
      - end_date   (YYYY-MM-DD)
    Если параметры не заданы:
      - для monthly trend: последние 12 месяцев
      - для daily/interactions: последние 30 дней
    """
    # parse dates
    sd = request.args.get('start_date')
    ed = request.args.get('end_date')

    # defaults
    now = datetime.utcnow()
    default_start_months = now - timedelta(days=365)  # 12 months
    default_start_days = now - timedelta(days=30)

    # We'll use start_date for monthly trends default to 12 months back if not provided,
    # and end_date default to now.
    try:
        if sd:
            start_date = datetime.strptime(sd, '%Y-%m-%d')
        else:
            start_date = default_start_months
        if ed:
            end_date = datetime.strptime(ed, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            end_date = now
    except ValueError:
        return jsonify({'error': 'invalid date format, use YYYY-MM-DD'}), 400

    # Aggregate metrics
    # Total debtors (all-time)
    total_debtors_all = db.session.query(func.count(Debtor.id)).scalar() or 0
    total_debt_all = float(db.session.query(func.coalesce(func.sum(Debtor.debt_amount), 0)).scalar() or 0.0)

    # Total debtors added in period
    total_debtors_period = db.session.query(func.count(Debtor.id)).filter(Debtor.created_at >= start_date, Debtor.created_at <= end_date).scalar() or 0
    total_debt_period = float(db.session.query(func.coalesce(func.sum(Debtor.debt_amount), 0)).filter(Debtor.created_at >= start_date, Debtor.created_at <= end_date).scalar() or 0.0)

    # By status (as of now, counting all debtors grouped by status)
    by_status_q = db.session.query(Debtor.status, func.count(Debtor.id)).group_by(Debtor.status).all()
    by_status = [{'status': s or '', 'count': int(c)} for s, c in by_status_q]

    # Debt by bank (top 20 by debt sum)
    debt_by_bank_q = (
        db.session.query(
            Debtor.bank_name,
            func.coalesce(func.sum(Debtor.debt_amount), 0).label('total'),
            func.count(Debtor.id).label('count')
        )
        .group_by(Debtor.bank_name)
        .order_by(desc('total'))
        .limit(20)
        .all()
    )
    debt_by_bank = [{'bank': b or 'Не указан', 'total_debt': float(total), 'count': int(cnt)} for b, total, cnt in debt_by_bank_q]

    # New debtors by month in the selected (or default) range - monthly buckets
    # Use SQLite strftime('%Y-%m', ...) for portability with sqlite; for Postgres use date_trunc.
    months_q = (
        db.session.query(func.strftime('%Y-%m', Debtor.created_at).label('month'), func.count(Debtor.id))
        .filter(Debtor.created_at >= start_date, Debtor.created_at <= end_date)
        .group_by('month')
        .order_by('month')
        .all()
    )
    new_debtors_month = [{'month': m, 'count': int(c)} for m, c in months_q]

    # Interactions per day in the period (grouped by day)
    interactions_q = (
        db.session.query(func.strftime('%Y-%m-%d', Interaction.date_time).label('day'), func.count(Interaction.id))
        .filter(Interaction.date_time >= start_date, Interaction.date_time <= end_date)
        .group_by('day')
        .order_by('day')
        .all()
    )
    interactions_by_day = [{'date': d, 'count': int(c)} for d, c in interactions_q]

    # Top debtors by amount (global, not limited by period) - top 20
    top_q = (
        db.session.query(Debtor.id, Debtor.full_name, Debtor.debt_amount, Debtor.bank_name, Debtor.status)
        .order_by(desc(Debtor.debt_amount))
        .limit(20)
        .all()
    )
    top_debtors = [
        {'id': d.id, 'full_name': d.full_name, 'debt_amount': float(d.debt_amount), 'bank': d.bank_name, 'status': d.status}
        for d in top_q
    ]

    # Interactions by user in period
    interactions_user_q = (
        db.session.query(User.id, User.username, func.count(Interaction.id).label('cnt'))
        .join(Interaction, Interaction.user_id == User.id)
        .filter(Interaction.date_time >= start_date, Interaction.date_time <= end_date)
        .group_by(User.id, User.username)
        .order_by(desc('cnt'))
        .all()
    )
    interactions_by_user = [{'user_id': r[0], 'username': r[1], 'count': int(r[2])} for r in interactions_user_q]

    # KPIs
    avg_debt = db.session.query(func.avg(Debtor.debt_amount)).scalar() or 0.0
    # median in python
    debt_values = [float(v[0]) for v in db.session.query(Debtor.debt_amount).filter(Debtor.debt_amount != None).all()]
    median_debt = 0.0
    if debt_values:
        debt_values_sorted = sorted(debt_values)
        n = len(debt_values_sorted)
        mid = n // 2
        if n % 2 == 1:
            median_debt = debt_values_sorted[mid]
        else:
            median_debt = (debt_values_sorted[mid-1] + debt_values_sorted[mid]) / 2.0

    total_all = total_debtors_all or 1
    paid_count = db.session.query(func.count(Debtor.id)).filter(Debtor.status == 'paid').scalar() or 0
    recovery_rate = float(paid_count) / float(total_all) if total_all else 0.0

    payload = {
        'total_debtors_all': int(total_debtors_all),
        'total_debt_all': float(total_debt_all),
        'total_debtors_period': int(total_debtors_period),
        'total_debt_period': float(total_debt_period),
        'by_status': by_status,
        'debt_by_bank': debt_by_bank,
        'new_debtors_month': new_debtors_month,
        'interactions_by_day': interactions_by_day,
        'top_debtors': top_debtors,
        'interactions_by_user': interactions_by_user,
        'kpis': {
            'avg_debt': float(avg_debt) if avg_debt is not None else 0.0,
            'median_debt': float(median_debt),
            'recovery_rate': float(recovery_rate)
        },
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d')
    }
    return jsonify(payload)


@bp.route('/analytics/export')
@login_required
@editor_required
def analytics_export():
    """
    Export selected analytics slices to CSV.
    Query params:
      - start_date, end_date (YYYY-MM-DD) to limit interactions and new_debtors if desired
      - what=top_debtors|interactions_by_user|summary (default top_debtors)
    Returns CSV as attachment.
    """
    what = request.args.get('what', 'top_debtors')
    sd = request.args.get('start_date')
    ed = request.args.get('end_date')
    try:
        if sd:
            start_date = datetime.strptime(sd, '%Y-%m-%d')
        else:
            start_date = datetime.utcnow() - timedelta(days=365)
        if ed:
            end_date = datetime.strptime(ed, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            end_date = datetime.utcnow()
    except ValueError:
        return "Invalid dates", 400

    output = io.StringIO()
    writer = csv.writer(output)

    if what == 'interactions_by_user':
        # header
        writer.writerow(['User ID', 'Username', 'Interactions Count'])
        rows = (
            db.session.query(User.id, User.username, func.count(Interaction.id).label('cnt'))
            .join(Interaction, Interaction.user_id == User.id)
            .filter(Interaction.date_time >= start_date, Interaction.date_time <= end_date)
            .group_by(User.id, User.username)
            .order_by(desc('cnt'))
            .all()
        )
        for r in rows:
            writer.writerow([r[0], r[1], int(r[2])])
        filename = f'analytics_interactions_by_user_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.csv'
    else:
        # default: top_debtors
        writer.writerow(['Debtor ID', 'Full name', 'Debt amount', 'Bank', 'Status'])
        rows = (
            db.session.query(Debtor.id, Debtor.full_name, Debtor.debt_amount, Debtor.bank_name, Debtor.status)
            .order_by(desc(Debtor.debt_amount))
            .limit(100)
            .all()
        )
        for r in rows:
            writer.writerow([r[0], r[1], float(r[2]) if r[2] is not None else 0.0, r[3], r[4]])
        filename = f'analytics_top_debtors_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.csv'

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8')), as_attachment=True, download_name=filename, mimetype='text/csv')
# --- End analytics block ---