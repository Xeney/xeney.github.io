import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from app import db
from app.models import Debtor

def import_excel(filepath):
    """Import debtors from Excel file"""
    stats = {'created': 0, 'updated': 0, 'errors': []}
    
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        # Check headers
        headers = [cell.value for cell in sheet[1]]
        expected_headers = ['ФИО', 'Дата рождения', 'Паспорт', 'Номер договора', 'Сумма долга', 'Банк']
        
        if headers != expected_headers:
            stats['errors'].append(f'Неверные заголовки. Ожидалось: {expected_headers}, получено: {headers}')
            return stats
        
        # Process rows
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
                
            try:
                full_name, birth_date_str, passport, contract, amount, bank = row
                
                # Validate required fields
                if not all([full_name, birth_date_str, passport, contract, amount, bank]):
                    stats['errors'].append(f'Строка {row_num}: Все поля обязательны для заполнения')
                    continue
                
                # Parse date
                try:
                    if isinstance(birth_date_str, datetime):
                        birth_date = birth_date_str.date()
                    elif isinstance(birth_date_str, str):
                        # Try different date formats
                        for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
                            try:
                                birth_date = datetime.strptime(birth_date_str, fmt).date()
                                break
                            except ValueError:
                                continue
                        else:
                            stats['errors'].append(f'Строка {row_num}: Неверный формат даты: {birth_date_str}')
                            continue
                    else:
                        stats['errors'].append(f'Строка {row_num}: Неверный тип данных для даты')
                        continue
                except Exception as e:
                    stats['errors'].append(f'Строка {row_num}: Ошибка парсинга даты: {str(e)}')
                    continue
                
                # Parse amount
                try:
                    if isinstance(amount, (int, float)):
                        debt_amount = float(amount)
                    else:
                        # Remove spaces and replace comma with dot
                        amount_str = str(amount).replace(' ', '').replace(',', '.')
                        debt_amount = float(amount_str)
                except (ValueError, TypeError):
                    stats['errors'].append(f'Строка {row_num}: Неверный формат суммы: {amount}')
                    continue
                
                # Check if debtor exists
                existing_debtor = Debtor.query.filter_by(contract_number=contract).first()
                
                if existing_debtor:
                    # Update existing
                    existing_debtor.full_name = full_name
                    existing_debtor.birth_date = birth_date
                    existing_debtor.passport_data = passport
                    existing_debtor.debt_amount = debt_amount
                    existing_debtor.bank_name = bank
                    stats['updated'] += 1
                else:
                    # Create new
                    debtor = Debtor(
                        full_name=full_name,
                        birth_date=birth_date,
                        passport_data=passport,
                        contract_number=contract,
                        debt_amount=debt_amount,
                        bank_name=bank
                    )
                    db.session.add(debtor)
                    stats['created'] += 1
                
            except Exception as e:
                stats['errors'].append(f'Строка {row_num}: Ошибка обработки данных: {str(e)}')
                continue
        
        db.session.commit()
        
    except Exception as e:
        stats['errors'].append(f'Ошибка чтения файла: {str(e)}')
    
    return stats